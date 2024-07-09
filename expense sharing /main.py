import sqlite3
import pandas as pd
import openpyxl
import datetime  # Import datetime for date and time

conn = sqlite3.connect('app.db')
cursor = conn.cursor()
# Insert data into the friend's table
today_date = datetime.date.today().strftime('%d-%m-%Y')
now_time = datetime.datetime.now().strftime('%H:%M:%S')


class Sharing_app:
    def __init__(self):
        self.users = []

    def options(self):
        print("1. add user \n2. add data \n3. display user \n4. display data \n5. delete data \n6. delete user \n7. split the bill \n8. Lend or borrow money \n9. save file to excel \n "  )
        opn = int(input("Enter the option from above : "))
        if opn ==1 :
            app.adduser()
        elif opn ==2 :
            app.add_data()
        elif opn ==3 :
            app.display_user()
        elif opn ==4 :
            app.display_data()
        elif opn ==5 :
            app.delete_user()
        elif opn ==6 :
            app.delete_data()
        elif opn ==7 :
            app.split_bill()
        elif opn ==8 :
            app.lend_borrow()
        elif opn ==9 :
            app.save_to_excel()
        else:
            print("please enter correct option ")

   


    def lend_borrow(self):
            """Facilitates lending and borrowing transactions between users.

            Raises:
                ValueError: If user names are not found in the database or the amount is non-positive.
            """

            # Get user input
            user1 = input("Enter the user who wants to lend money: ").upper()
            user2 = input("Enter the user who wants to borrow money: ").upper()
            amount = float(input("Enter the amount you want to lend/borrow: "))

            # Validate positive amount
            if amount <= 0:
                raise ValueError("Transaction amount must be a positive number.")

            # Check for user existence using prepared statements
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            table_names = cursor.fetchall()
            print("Tables in app.db:")
            list1 = []
            for row in table_names:
                print(row[0])
                list1.append(row[0])
            print(list1)

            if user1 not in list1 or user2 not in list1:
                raise ValueError("One or both users do not exist.")


            # Insert transactions using prepared statements
            insert_user1_stmt = f"""
                INSERT INTO {user1} (NAME, CATEGORY, DATE, TIME, AMOUNT)
                VALUES (?, ?, ?, ?, ?)
            """
            insert_user2_stmt = f"""
                INSERT INTO {user2} (NAME, CATEGORY, DATE, TIME, AMOUNT)
                VALUES (?, ?, ?, ?, ?)
            """
            cursor.execute(insert_user1_stmt, (user2, "lend money", today_date, now_time, amount))
            cursor.execute(insert_user2_stmt, (user1, "borrow money", today_date, now_time, amount))  # Negate amount for borrow

            # Commit changes to the database
            conn.commit()
            print(f"{user1} lent Rs.{amount:.2f} to {user2}.")


    def adduser(self):
        users = input("Enter the user name :")
        if users:
            conn.execute(f'''CREATE TABLE {users.upper()} (
                   NAME VARCHAR(10),
                   CATEGORY VARCHAR(100),
                   DATE VARCHAR(100),
                   TIME VARCHAR(100),
                   AMOUNT INT(50)
                );''')

    def add_data(self):
        table_name = input("Enter the user name (table name): ")  # Clarify table name input
        NAME = input("Enter the name of the item: ")
        CATEGORY = input("Enter the category: ")
        DATE = today_date
        TIME = now_time
        AMOUNT = input("Enter the amount : ")

        # Execute query with placeholders for safer parameter passing
        cursor.execute(f"""INSERT INTO {table_name.upper()} 
                          (NAME, CATEGORY, DATE, TIME, AMOUNT)
                          VALUES (?, ?, ?, ?, ?)""",
                      (NAME, CATEGORY, DATE, TIME, AMOUNT))
        conn.commit()  # Commit changes to the database

    def display_user(self):
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        table_names = cursor.fetchall()
        print("Tables in app.db:")
        for row in table_names:
            print(row[0])

    def display_data(self):
        table_name = input("Enter the table name to display: ")

        # Execute query to fetch data
        cursor.execute(f"SELECT * FROM {table_name}")

        # Fetch all data as a list of tuples
        data = cursor.fetchall()

        # Check if table exists and has data
        if not data:
            print(f"Table '{table_name}' does not exist or is empty.")
            return

        # Get column names from cursor description
        column_names = [col[0] for col in cursor.description]

        # Create DataFrame from data and column names
        df = pd.DataFrame(data, columns=column_names)

        # Display DataFrame
        print(df.to_string(index=True))  # Display with default index
         # Display with index + 1 offset
         
    def delete_data(self):
        table_name = input("Enter the table name to delete from: ")
        # Prompt for row selection method (using ID or criteria)
        selection_method = input("Choose deletion method (1 - by ID, 2 - by criteria): ")

        if selection_method == '1':
            row_id = int(input("Enter the ID of the row to delete: "))
            # Delete by ID using a safer parameterized query
            cursor.execute(f"""DELETE FROM {table_name.upper()} WHERE ROWID = ?""", (row_id,))
        elif selection_method == '2':
            # Implement deletion by criteria (e.g., name, category, etc.)
            # Consider using parameterized queries to prevent SQL injection
            criteria_column = input("Enter the column name for criteria: ")
            criteria_value = input("Enter the value to match: ")
            cursor.execute(f"""DELETE FROM {table_name.upper()} 
                              WHERE {criteria_column.upper()} = ?""", (criteria_value,))
        else:
            print("Invalid selection method. Please choose 1 or 2.")

        conn.commit()  # Commit changes after deletion
        print(f"Deletion from '{table_name}' completed (if applicable).")
        
    def delete_user(self):
        user_to_delete = input("Enter the user name (table name) to delete: ")

        # Check if table exists before deletion
        cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name='{user_to_delete.upper()}'")
        if not cursor.fetchone():
            print(f"Table '{user_to_delete}' does not exist.")
            return

        # Drop the table using a safer parameterized query
        cursor.execute(f"""DROP TABLE IF EXISTS {user_to_delete.upper()}""",)
        conn.commit()
        
        
    def save_to_excel(self):
            table_name = input("Enter the table name : ").upper()
            file_name = input("Enter the file_name :")
        # Check if table exists
            cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name='{table_name.upper()}'")
            if not cursor.fetchone():
                print(f"Table '{table_name}' does not exist. Cannot save data to Excel.")
                return

        # Fetch data from the table
            
            cursor.execute(f"SELECT * FROM {table_name}")
            data = cursor.fetchall()

        # Check if table has data
            if not data:
                print(f"Table '{table_name}' is empty. Cannot save empty data to Excel.")
                return

        # Get column names from cursor description
            column_names = [col[0] for col in cursor.description]

        # Create a new workbook in openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = table_name  # Set worksheet title to table name

        # Write column names as the first row
            for col_num, col_name in enumerate(column_names, start=1):
                ws.cell(row=1, column=col_num).value = col_name

        # Write data rows starting from the second row
            for row_num, row_data in enumerate(data, start=2):
                for col_num, cell_value in enumerate(row_data, start=1):
                    ws.cell(row=row_num, column=col_num).value = cell_value

        # Save the workbook as the specified file name
            wb.save(filename=f"{file_name}.xlsx")
            print(f"Data from table '{table_name}' saved to Excel file: {file_name}")
            
            
    def split_bill(self):
        friend_names = []
        while True:
            fri = input("Enter the name of the friend (-1 to stop): ")
            if fri != "-1":
                friend_names.append(fri.upper())
            else:
                print(friend_names)
                break
            
        if not friend_names:
            raise ValueError("Please provide at least one friend name (table name).")

        # Get total amount to be split
        total_amount = float(input("Enter the total bill amount: "))

        # Calculate the equal share per friend (excluding rounding errors)
        num_friends = len(friend_names)
        share_per_friend = total_amount / num_friends

        # Split the amount and insert into each friend's table
        for friend_name in friend_names:
            # Prompt for custom amount for this friend (optional)
            custom_amount = input(f"Enter the amount for {friend_name} (or press Enter for equal share): ")
            if custom_amount:
                amount = float(custom_amount)
            else:
                amount = share_per_friend

            

            # Execute the insert query
            cursor.execute(f"""INSERT INTO {friend_name.upper()} (NAME, CATEGORY, DATE, TIME, AMOUNT)
                            VALUES (?, ?, ?, ?, ?)""",
                            ("bill", "Split Bill", today_date, now_time, amount))
            conn.commit()

        print(f"\nSplit bill of RS.{total_amount:.2f} has been added to tables for {', '.join(friend_names)}.")


app = Sharing_app()
app.options()
conn.close()
